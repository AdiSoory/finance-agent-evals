#!/usr/bin/env python3
"""
Workbook Integrity Scorer (Dimension D7) -- finance-agent-evals
Version: 0.2.0

Audits an AI-generated Excel model for structural integrity:
  1. Formula-vs-hardcode ratio per sheet
  2. Orphaned assumptions tab (inputs no other sheet references)
  3. STRUCTURAL constants embedded inside formulas (rates, percentages,
     dollar amounts that should live on an assumptions tab) -- penalized
  4. Conventional constants (date criteria, /12 month conversions, etc.)
     -- reported for information, NOT penalized
  5. Hardcoded non-zero numbers on output sheets

Changes in 0.2.0:
  - Quoted string literals (e.g. "2025-06") are stripped before constant
    scanning: date strings no longer flagged
  - Constants classified structural vs conventional; only structural
    constants affect the score
  - Whitelist: calendar/structural integers (1-4, 12, 100, years 1900-2100)
  - Deduplication per (sheet, cell, constant)
  - Literal zeros on output sheets ignored (placeholders, not results)
  - Graceful exit on broken pipe; --json flag; version stamped in report

Usage:
    pip install openpyxl
    python score_workbook.py path/to/model.xlsx [--json]

Score: 0-4. See README.
"""

import json
import re
import sys
from collections import defaultdict

try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit("openpyxl is required: pip install openpyxl")

VERSION = "0.2.0"

# Sheets whose names suggest raw data or inputs (hardcodes expected there)
INPUT_SHEET_PAT = re.compile(r"assumption|input|ledger|data|readme|raw", re.I)
DATA_SHEET_PAT = re.compile(r"ledger|data|readme|raw", re.I)

QUOTED_STRING = re.compile(r'"[^"]*"')
CROSS_SHEET_REF = re.compile(r"'([^']+)'!|(\b[A-Za-z_][\w .&]*)!")
# numeric tokens not part of a cell reference or function name
NUM_TOKEN = re.compile(r"(?<![A-Za-z$:!.\d])(\d+(?:\.\d+)?)(?![A-Za-z0-9(.])")

# Integers that are conventional in finance formulas, not smuggled drivers
CONVENTIONAL_INTS = {1, 2, 3, 4, 12, 100, 365, 360, 30, 90, 1000, 1000000}


def classify_constant(tok: str) -> str:
    """Return 'structural' (penalized) or 'conventional' (informational)."""
    val = float(tok)
    if val == 0:
        return "conventional"
    if val.is_integer():
        iv = int(val)
        if iv in CONVENTIONAL_INTS:
            return "conventional"
        if 1900 <= iv <= 2100:          # year criteria in SUMIFS etc.
            return "conventional"
        return "structural"              # e.g. 120000, 1500000
    # Non-integer decimals are almost always rates/shares: 0.08, 0.22, 0.5
    return "structural"


def audit(path: str) -> dict:
    wb = load_workbook(path)
    report = {
        "scorer_version": VERSION,
        "file": path,
        "sheets": {},
        "assumption_sheets": [],
        "orphaned_assumption_sheets": [],
        "structural_constants_in_formulas": [],
        "conventional_constants_in_formulas": [],
        "hardcoded_outputs": [],
    }

    xref = defaultdict(int)   # sheet -> formula references from OTHER sheets
    seen = set()              # dedupe (sheet, cell, token)

    for sn in wb.sheetnames:
        ws = wb[sn]
        n_formula, n_hardcode = 0, 0
        for row in ws.iter_rows():
            for c in row:
                v = c.value
                if v is None:
                    continue
                if isinstance(v, str) and v.startswith("="):
                    n_formula += 1
                    for m in CROSS_SHEET_REF.finditer(v):
                        ref = (m.group(1) or m.group(2)).strip()
                        if ref and ref != sn:
                            xref[ref] += 1
                    # strip quoted strings BEFORE scanning for constants,
                    # so "2025-06" and similar date criteria are invisible
                    stripped = QUOTED_STRING.sub('""', v)
                    for m in NUM_TOKEN.finditer(stripped):
                        tok = m.group(1)
                        key = (sn, c.coordinate, tok)
                        if key in seen:
                            continue
                        seen.add(key)
                        entry = {"sheet": sn, "cell": c.coordinate,
                                 "constant": tok, "formula": v[:90]}
                        kind = classify_constant(tok)
                        report[f"{kind}_constants_in_formulas"].append(entry)
                elif isinstance(v, (int, float)):
                    n_hardcode += 1
                    # zeros are placeholders, not smuggled results
                    if v != 0 and not INPUT_SHEET_PAT.search(sn):
                        report["hardcoded_outputs"].append(
                            {"sheet": sn, "cell": c.coordinate, "value": v})
        report["sheets"][sn] = {"formulas": n_formula,
                                "hardcoded_numbers": n_hardcode}
        if INPUT_SHEET_PAT.search(sn) and not DATA_SHEET_PAT.search(sn):
            report["assumption_sheets"].append(sn)

    for sn in report["assumption_sheets"]:
        if xref.get(sn, 0) == 0:
            report["orphaned_assumption_sheets"].append(sn)

    report["cross_sheet_reference_counts"] = dict(xref)
    return report


def score(report: dict) -> int:
    """0-4 integrity score. Only structural findings are penalized."""
    s = 4
    if report["orphaned_assumption_sheets"]:
        s -= 2  # assumptions tab is decorative: model is static
    if len(report["hardcoded_outputs"]) > 20:
        s -= 1
    if len(report["structural_constants_in_formulas"]) > 3:
        s -= 1
    return max(s, 0)


def render(report: dict) -> str:
    lines = []
    p = lines.append
    p(f"\n=== Workbook Integrity Report (scorer v{VERSION}) ===")
    p(f"File: {report['file']}\n")
    p(f"{'Sheet':<32}{'Formulas':>10}{'Hardcodes':>10}")
    for sn, d in report["sheets"].items():
        p(f"{sn:<32}{d['formulas']:>10}{d['hardcoded_numbers']:>10}")

    p("\nAssumption/input sheets detected: "
      + (", ".join(report["assumption_sheets"]) or "none"))
    if report["orphaned_assumption_sheets"]:
        p("!! ORPHANED assumption sheets (no other sheet references them): "
          + ", ".join(report["orphaned_assumption_sheets"]))
        p("   -> Changing these inputs changes NOTHING. "
          "This is a report, not a model.")
    elif report["assumption_sheets"]:
        p("Assumption sheets are referenced by downstream formulas. OK.")

    sc = report["structural_constants_in_formulas"]
    p(f"\nSTRUCTURAL constants inside formulas (penalized): {len(sc)}")
    p("  (rates, shares, or dollar amounts that belong on an assumptions tab)")
    for item in sc[:10]:
        p(f"  {item['sheet']}!{item['cell']}: {item['formula']}"
          f"  (constant {item['constant']})")
    if len(sc) > 10:
        p(f"  ... and {len(sc)-10} more")

    cc = report["conventional_constants_in_formulas"]
    p(f"\nConventional constants (informational, not penalized): {len(cc)}")
    p("  (calendar integers, year criteria, unit conversions)")

    h = report["hardcoded_outputs"]
    p(f"\nHardcoded NON-ZERO numbers on output sheets: {len(h)}")
    for item in h[:10]:
        p(f"  {item['sheet']}!{item['cell']} = {item['value']}")
    if len(h) > 10:
        p(f"  ... and {len(h)-10} more")

    p(f"\nD7 WORKBOOK INTEGRITY SCORE: {score(report)} / 4\n")
    return "\n".join(lines)


def main():
    args = [a for a in sys.argv[1:] if a != "--json"]
    as_json = "--json" in sys.argv
    if len(args) != 1:
        sys.exit(__doc__)
    report = audit(args[0])
    report["score"] = score(report)
    try:
        if as_json:
            print(json.dumps(report, indent=2))
        else:
            print(render(report))
    except BrokenPipeError:
        sys.stderr.close()


if __name__ == "__main__":
    main()
